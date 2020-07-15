import pandas as pd
import sys
import os.path as path
import pymysql
import numpy as np
from openpyxl import Workbook, cell
from openpyxl.styles import Font
import os
import shutil


def styled_dept_floated_headers():
    data = ("Roll No", "Email", "Name", "Department", "Course Alloted")
    for c in data:
        c = cell.Cell(ws, column="A", row=1, value=c)
        c.font = Font(bold=True)
        yield c
    return data


def styled_unallocated_students_headers(worksheet):
    data = ("Roll No", "Email", "Name", "Department")
    for c in data:
        c = cell.Cell(worksheet, column="A", row=1, value=c)
        c.font = Font(bold=True)
        yield c
    return data


def style_error(student_data, worksheet):
    for c in student_data:
        c = cell.Cell(worksheet, column="A", row=1, value=c)
        c.font = Font(bold=True, color="ff0000")
        yield c
    return student_data


def styled_dept_student_headers(worksheet):
    data = ("Roll No", "Email", "Name",
            "Course Alloted", "Offering Department")
    for c in data:
        c = cell.Cell(worksheet, column="A", row=1, value=c)
        c.font = Font(bold=True)
        yield c
    return data


mapper = {
    "sem": 0,
    "year": 1,
    "student_pref_table": 2,
    "student_course_table": 3,
    "course_allocate_info_table": 4,
    "course_table": 5,
    "pref_percent_table": 6,
    "pref_student_alloted_table": 7,
    "course_app_dept_table": 8,
    "no_of_preferences": 9,
    "host": 10,
    "username": 11,
    "password": 12,
    "dbname": 13,
    "program": 14,
    "form_id": 15,
    "save_location": 16
}
argument = list(map(str.strip, sys.argv[1].strip('[]').split(',')))
mydb = pymysql.connect(
    host=argument[mapper['host']],
    user=argument[mapper['username']],
    passwd=argument[mapper['password']],
    database=argument[mapper['dbname']]
)
mycursor = mydb.cursor()


try:
    form_course_ids_query = "select GROUP_CONCAT(ct.name SEPARATOR '-') from form_course_category_map fccm INNER join course_types ct on fccm.course_type_id=ct.id where fccm.form_id=%s"
    mycursor.execute(form_course_ids_query, (argument[mapper['form_id']],))

    course_types_string = mycursor.fetchall()[0][0]

    folder_name = course_types_string + "-sem-" + \
        argument[mapper["sem"]] + "-year-" + \
        argument[mapper["year"]] + "-allocation"

    folder_path = argument[mapper["save_location"]] + folder_name

    if (path.isdir(folder_path)):
        for root, dirs, files in os.walk(folder_path):
            for f in files:
                os.unlink(os.path.join(root, f))
            for d in dirs:
                shutil.rmtree(os.path.join(root, d))
    else:
        os.mkdir(folder_path)
    dept_courses_folder = folder_path + "/dept_courses_allocation"
    dept_students_folder = folder_path+"/dept_students_allocation"
    os.mkdir(dept_courses_folder)
    os.mkdir(dept_students_folder)

    dept_applicable_workbooks = {}

    dept_student_workbooks = {}

    student_alloted_dept_applicable_query = "SELECT s.rollno, cfd.cid,d.dept_name as floating_dept, c.cname," + \
        " s.email_id, concat(s.fname, ' ', s.mname, ' ', s.lname) as name , d2.dept_name as student_dept" + \
        " FROM "+argument[mapper['student_course_table']]+" sca " + \
        "inner join course_floating_dept cfd " +\
        "inner join student s " +\
        "inner join department d " + \
        "inner join department d2 " +\
        "inner join "+argument[mapper['course_table']]+" c " +\
        "on " +\
        "s.email_id=sca.email_id and sca.cid=cfd.cid " +\
        "and sca.sem = cfd.sem and sca.year=cfd.year " + \
        "and s.dept_id = d2.dept_id " +\
        "and cfd.dept_id = d.dept_id and sca.cid = c.cid and sca.sem = c.sem and sca.year = c.year ORDER BY d.dept_name,d2.dept_name,s.rollno"

    mycursor.execute(student_alloted_dept_applicable_query)

    for result_row in mycursor.fetchall():
        wb = None
        wb_student = None
        ws = None
        ws_student = None

        if (result_row[2] in dept_applicable_workbooks):  # department wise workbook
            wb = dept_applicable_workbooks[result_row[2]]
        else:
            wb = Workbook()
            dept_applicable_workbooks[result_row[2]] = wb
            ws = wb.active
            ws.title = result_row[1]
            ws.append(styled_dept_floated_headers())

        if (result_row[1] in wb.sheetnames):
            ws = wb[result_row[1]]
        else:
            ws = wb.create_sheet(result_row[1])
            ws.append(styled_dept_floated_headers())

        data = (int(result_row[0]), result_row[4], result_row[5],
                result_row[6], result_row[3])
        ws.append(data)

    for key in dept_applicable_workbooks.keys():
        dept_applicable_workbooks[key].save(
            dept_courses_folder+"/" + key + ".xlsx")

    department_students_info_query = "SELECT s.rollno,d2.dept_name as student_dept,concat(s.fname, ' ', s.mname, ' ', s.lname) as name ,sp.email_id, " +\
        "ifnull(c.cname, 'No Course Alloted') as course_alloted, ifnull(d.dept_name, 'NA') as floating_dept " +\
        "FROM(`"+argument[mapper["student_pref_table"]]+"` as sp INNER JOIN student s inner join department d2 on sp.email_id=s.email_id and s.dept_id=d2.dept_id)" +\
        "left join(`" + argument[mapper['student_course_table']] + "` as sca INNER JOIN `"+argument[mapper["course_table"]]+"` c " +\
        "INNER JOIN course_floating_dept cfd INNER JOIN department D " +\
        "on sca.cid=c.cid and sca.sem=c.sem and sca.year=c.year and sca.cid=cfd.cid and sca.sem=cfd.sem and " +\
        "sca.year=cfd.year and cfd.dept_id=d.dept_id) on sp.email_id = sca.email_id and c.cid = sca.cid order by s.rollno"

    mycursor.execute(department_students_info_query)
    # print(mycursor.fetchall())
    for result_row in mycursor.fetchall():
        # print(result_row[1])
        wb_student = None
        ws_student = None
        if (result_row[1] in dept_student_workbooks):
            wb_student = dept_student_workbooks[result_row[1]]
            ws_student = wb_student.active
        else:
            wb_student = Workbook()
            dept_student_workbooks[result_row[1]] = wb_student
            ws_student = wb_student.active
            ws_student.append(styled_dept_student_headers(ws_student))

        student_data = (int(result_row[0]), result_row[3], result_row[2],
                        result_row[4], result_row[5])
        if (result_row[4] == "No Course Alloted"):
            # print("a")
            student_data = style_error(student_data, ws_student)

        ws_student.append(student_data)

    for key in dept_student_workbooks.keys():
        dept_student_workbooks[key].save(
            dept_students_folder + "/" + key + ".xlsx")

    unallocated_students_query = "SELECT p.email_id, p.rollno, CONCAT(fname, ' ', mname, ' ', lname) as fullname, dept_name, " +\
        "p.timestamp FROM "+argument[mapper['student_pref_table']]+" p INNER JOIN student s " +\
        "INNER JOIN department d ON p.email_id = s.email_id AND s.dept_id = d.dept_id WHERE allocate_status = '0' order by dept_name,p.rollno"

    mycursor.execute(unallocated_students_query)
    unallocated_workbook = Workbook()
    unallocated_worksheet = unallocated_workbook.active
    unallocated_worksheet.append(
        styled_unallocated_students_headers(unallocated_worksheet))
    for x in mycursor.fetchall():
        unallocated_worksheet.append((x[1], x[0], x[2], x[3]))
    unallocated_workbook.save(folder_path+"/unallocated_students.xlsx")

    course_query = "SELECT cid,sem,year,min,max FROM " + \
        argument[mapper['course_table']]+" WHERE sem='" + \
        argument[mapper['sem']]+"' and year='"+argument[mapper['year']]+"'"
    mycursor.execute(course_query)

    myresult = mycursor.fetchall()
    courses = {}
    update_course_query = "update course set min={} and max={} where cid='{}' and sem={} and year='{}' ;"
    for x in myresult:
        try:
            mycursor.execute(update_course_query.format(
                int(x[3]), x[4], x[0], x[1], x[2]))
        except Exception as e:
            print("error+", e)
            sys.exit(0)

    copy_student_alloted_query = "insert into student_course_alloted (email_id,cid,sem,year,course_type_id) SELECT email_id,cid,sem,year, (select course_type_id from course where cid=sc.cid and sem=sc.sem and year=sc.year) as course_type_id FROM " + \
        argument[mapper['student_course_table']] + " as sc"

    copy_student_alloted_to_grade_table = "insert into student_courses_grade (email_id, cid, sem, year, course_type_id) SELECT email_id, cid, sem, year, (select course_type_id from course where cid=sc.cid and sem=sc.sem and year=sc.year) as course_type_id FROM " + \
        argument[mapper['student_course_table']] + " as sc"
    update_in_student_preferences = "update student_preferences set allocate_status=1 where form_id='" + argument[mapper["form_id"]] + "' " + \
        "and email_id in (select email_id from " + \
        argument[mapper["student_course_table"]]+")"
    try:
        mycursor.execute(copy_student_alloted_query)
        mycursor.execute(copy_student_alloted_to_grade_table)
        mycursor.execute(update_in_student_preferences)
    except Exception as e:
        print("error+", e)
        mycursor.close()
        sys.exit(0)
    shutil.make_archive(folder_path, 'zip', folder_path)

    print("done+"+folder_path)

except Exception as e:
    print("error+", e)
    mycursor.close()
    sys.exit(0)

finally:
    mydb.commit()
    mycursor.close()
